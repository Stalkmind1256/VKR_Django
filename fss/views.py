import json

from django.contrib.auth.hashers import make_password
from django.shortcuts import render, get_object_or_404, redirect
from django.utils.dateparse import parse_date
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from openpyxl.workbook import Workbook
from django.contrib.auth import get_user_model
from django.db.models import Avg
from django.db.models import Q
from .models import Category, Divisions, Suggestion, Status, Notification, Comment, CustomUser, SuggestionRating
from .forms import SuggestionForm, CommentForm, CustomUserEditForm
from django.contrib.auth.forms import UserCreationForm
from django.db.models import Count
from .models import Suggestion
from django.utils.timezone import is_aware
import openpyxl
from django.views.decorators.http import require_POST
from .models import CustomUser
from .forms import CustomUserCreationForm
from .models import CustomUser

@login_required
def home(request):
    unread_count = request.user.notifications.filter(is_read=False).count()

    best_suggestions = Suggestion.objects.filter(
        status__name__in=['approved', 'completed']
    ).annotate(
    annotated_avg_rating=Avg('ratings__rating'),
    annotated_votes_count=Count('ratings')
).order_by('-date_create')[:5]

    user_ratings_qs = SuggestionRating.objects.filter(user=request.user, suggestion__in=best_suggestions)
    user_ratings = {r.suggestion_id: r.rating for r in user_ratings_qs}

    return render(request, 'fss/home.html', {
        'unread_count': unread_count,
        'best_suggestions': best_suggestions,
        'user_ratings': user_ratings,
    })

def register(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            username = form.cleaned_data.get('username')
            messages.success(request, f'Пользователь {username} успешно зарегистрирован')
            return redirect('login')
    else:
        form = UserCreationForm()
    return render(request, 'registration/register.html', {'form': form})


def suggestion_list(request):
    # Получаем фильтры из GET-запроса
    status_filter = request.GET.get('status')
    category_filter = request.GET.get('category')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Базовый запрос
    suggestions = Suggestion.objects.select_related('category', 'status').all().order_by('-id')

    # Применяем фильтры
    if category_filter:
        suggestions = suggestions.filter(category__name=category_filter)
    if status_filter:
        suggestions = suggestions.filter(status__name=status_filter)
    if start_date:
        suggestions = suggestions.filter(date_create__gte=parse_date(start_date))
    if end_date:
        suggestions = suggestions.filter(date_create__lte=parse_date(end_date))

    # Пагинация
    paginator = Paginator(suggestions, 10)
    page = request.GET.get('page')
    try:
        page_obj = paginator.page(page)
    except (PageNotAnInteger, EmptyPage):
        page_obj = paginator.page(1)

    return render(request, 'fss/suggestion_list.html', {
        'page_obj': page_obj,
        'status_filter': status_filter,
        'category_filter': category_filter,
        'start_date': start_date,
        'end_date': end_date,
    })


@login_required
def suggestion_detail(request, suggestion_id):
    suggestion = get_object_or_404(Suggestion, id=suggestion_id)
    comments_list = suggestion.comments.order_by('-created_at')

    paginator = Paginator(comments_list, 4)
    page = request.GET.get('page')
    comments = paginator.get_page(page)

    form = CommentForm()

    return render(request, 'fss/suggestion_detail.html', {
        'suggestion': suggestion,
        'comments': comments,
        'form': form,
    })


@login_required
def add_comment(request, suggestion_id):
    suggestion = get_object_or_404(Suggestion, id=suggestion_id)

    if request.method == 'POST':
        form = CommentForm(request.POST)
        if form.is_valid():
            comment = form.save(commit=False)
            comment.suggestion = suggestion
            comment.user = request.user
            comment.save()
            return redirect('suggestion_detail', suggestion_id=suggestion.id)

    return redirect('suggestion_detail', suggestion_id=suggestion.id)


@login_required
def create_suggestion(request):
    if request.method == 'POST':
        form = SuggestionForm(request.POST)
        if form.is_valid():
            suggestion = form.save(commit=False)
            suggestion.user = request.user
            suggestion.status = Status.objects.get(name='draft')
            # Подставляем подразделение из пользователя:
            suggestion.division = request.user.division
            suggestion.save()
            return redirect('my_suggestions')
    else:
        form = SuggestionForm()
    return render(request, 'fss/suggestion_form.html', {'form': form})


@login_required
def edit_suggestion(request, pk):
    suggestion = get_object_or_404(Suggestion, pk=pk, user=request.user)
    if suggestion.status.name != 'draft':
        return redirect('my_suggestions')

    if request.method == 'POST':
        form = SuggestionForm(request.POST, instance=suggestion)
        if form.is_valid():
            form.save()
            return redirect('my_suggestions')
    else:
        form = SuggestionForm(instance=suggestion)
    return render(request, 'fss/suggestion_form.html', {'form': form})


@login_required
def submit_suggestion(request, pk):
    suggestion = get_object_or_404(Suggestion, pk=pk, user=request.user)

    if suggestion.status.name == 'draft':
        if suggestion.can_change_status('submitted'):
            suggestion.status = Status.objects.get(name='submitted')
            suggestion.save()

            # ✅ Уведомление администраторам
            User = get_user_model()
            admins = User.objects.filter(is_superuser=True)

            for admin in admins:
                Notification.objects.create(
                    user=admin,
                    message=f"Новое предложение от {request.user.get_full_name() or request.user.username}: «{suggestion.title}»"
                )

            messages.success(request, "Статус успешно изменён на 'Отправлено'. Администратор уведомлён.")
        else:
            messages.error(request, "Переход из статуса 'Черновик' в 'Отправлено' не разрешён.")
    else:
        messages.error(request, "Статус предложения не 'Черновик', изменение невозможно.")

    return redirect('my_suggestions')


@login_required
def moderator_panel(request):
    suggestions = Suggestion.objects.select_related('user', 'status').order_by('-id')

    status_filter = request.GET.get('status')
    user_filter = request.GET.get('user')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if status_filter:
        suggestions = suggestions.filter(status__name=status_filter)
    if user_filter:
        suggestions = suggestions.filter(user__username__icontains=user_filter)
    if start_date:
        suggestions = suggestions.filter(date_create__gte=parse_date(start_date))
    if end_date:
        suggestions = suggestions.filter(date_create__lte=parse_date(end_date))

    paginator = Paginator(suggestions, 10)
    page = request.GET.get('page')
    suggestions_page = paginator.get_page(page)

    statuses = Status.objects.all()

    return render(request, 'fss/moderator_panel.html', {
        'suggestions': suggestions_page,
        'status_filter': status_filter,
        'user_filter': user_filter,
        'start_date': start_date,
        'end_date': end_date,
        'statuses': statuses,
    })


@login_required
def my_suggestions(request):
    user_suggestions_list = Suggestion.objects.filter(user=request.user).order_by('-id')
    paginator = Paginator(user_suggestions_list, 5)
    page = request.GET.get('page')
    suggestions = paginator.get_page(page)
    return render(request, 'fss/my_suggestions.html', {'suggestions': suggestions})


@login_required
def reject_suggestion(request):
    if request.method == 'POST':
        suggestion_id = request.POST.get('suggestion_id')
        reason = request.POST.get('reason')
        action = request.POST.get('action')  # новый статус

        try:
            suggestion = Suggestion.objects.get(id=suggestion_id)

            if suggestion.status.name == 'archived':
                return JsonResponse({'success': False, 'error': 'Нельзя изменить статус архивированного предложения'})

            if not suggestion.can_change_status(action):
                return JsonResponse({'success': False, 'error': f"Переход из '{suggestion.status.name}' в '{action}' запрещён"})

            status_obj = Status.objects.get(name=action)
            suggestion.status = status_obj
            suggestion.save()

            Comment.objects.create(
                suggestion=suggestion,
                user=request.user,
                text=reason
            )

            return JsonResponse({
                'success': True,
                'id': suggestion.id,
                'new_status': status_obj.get_name_display(),
                'status_class': get_status_class(status_obj.name),
            })
        except (Suggestion.DoesNotExist, Status.DoesNotExist) as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Invalid request'})


@login_required
def approve_suggestion(request):
    if request.method == "POST":
        suggestion_id = request.POST.get("suggestion_id")
        status_name = request.POST.get("status")
        comment = request.POST.get("comment", "")

        suggestion = get_object_or_404(Suggestion, id=suggestion_id)

        # Проверяем, можно ли перейти в нужный статус
        if not suggestion.can_change_status(status_name):
            return JsonResponse({
                "success": False,
                "error": f"Переход из статуса '{suggestion.status.name}' в '{status_name}' не разрешён."
            }, status=400)

        # Обновляем статус
        status = Status.objects.get(name=status_name)
        suggestion.status = status
        suggestion.save()

        # Добавляем комментарий, если есть
        if comment:
            Comment.objects.create(
                suggestion=suggestion,
                user=request.user,
                text=comment
            )

        # Уведомление пользователю
        Notification.objects.create(
            user=suggestion.user,
            message=f"Статус вашего предложения «{suggestion.title}» изменён на «{status.get_name_display()}»."
        )

        return JsonResponse({
            "success": True,
            "id": suggestion.id,
            "new_status": status.get_name_display(),
            "status_class": get_status_class(status.name),
        })

    return JsonResponse({"success": False, "error": "Метод не разрешен"}, status=405)



@login_required
def notifications_view(request):
    notifications_list = request.user.notifications.order_by('-created_at')
    paginator = Paginator(notifications_list, 10)  # по 15 уведомлений на страницу

    page_number = request.GET.get('page')
    try:
        notifications = paginator.page(page_number)
    except PageNotAnInteger:
        notifications = paginator.page(1)
    except EmptyPage:
        notifications = paginator.page(paginator.num_pages)

    context = {
        'notifications': notifications
    }
    return render(request, 'fss/notifications.html', context)

@login_required
def unread_notification_count(request):
    count = Notification.objects.filter(user=request.user, is_read=False).count()
    return JsonResponse({'unread_count': count})

def get_status_class(status_name):
    status_classes = {
        'draft': 'bg-secondary',
        'submitted': 'bg-primary',
        'rejected': 'bg-danger',
        'archived': 'bg-dark',
        'approved': 'bg-success',
        'preparing': 'bg-info',
        'in_progress': 'bg-warning',
        'completed': 'bg-success',
    }
    return status_classes.get(status_name, 'bg-warning')

def suggestions_stats_api(request):
    qs = Suggestion.objects.values('status__name').annotate(count=Count('id'))

    data = {item['status__name']: item['count'] for item in qs}

    return JsonResponse(data)

def stats(request):

    return render(request, 'fss/stats.html')

@login_required
def export_suggestions_csv(request):
    import csv
    from django.http import HttpResponse

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="suggestions.csv"'

    response.write('\ufeff')

    writer = csv.writer(response)
    writer.writerow(['ID', 'Пользователь', 'Заголовок', 'Статус', 'Дата'])

    for s in Suggestion.objects.all():
        writer.writerow([s.id, s.user.username, s.title, s.status.get_name_display(), s.date_create])

    return response

# Excel экспорт
@login_required
def export_suggestions_excel(request):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="suggestions.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Предложения"

    ws.append(['ID', 'Пользователь', 'Название', 'Статус', 'Дата создания'])

    for s in Suggestion.objects.all():
        date_naive = s.date_create.replace(tzinfo=None) if is_aware(s.date_create) else s.date_create
        ws.append([
            s.id,
            s.user.username,
            s.title,
            s.status.get_name_display(),
            date_naive,
        ])

    wb.save(response)
    return response


def import_users(request):
    if request.method == 'POST':
        excel_file = request.FILES['file']
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            last_name, first_name, patronymic, division_name, username, raw_password = row

            division, _ = Divisions.objects.get_or_create(name=division_name)

            if not CustomUser.objects.filter(username=username).exists():
                CustomUser.objects.create(
                    username=username,
                    first_name=first_name,
                    last_name=last_name,
                    password=make_password(raw_password),
                    division=division,
                    patronymic=patronymic
                )

        messages.success(request, "Пользователи успешно импортированы.")
        return redirect('import_users')

    return render(request, 'fss/import_users.html')

def user_management(request):
    query = request.GET.get("q", "")
    users = CustomUser.objects.all()

    if query:
        users = users.filter(
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(patronymic__icontains=query) |
            Q(division__name__icontains=query)
        )

    paginator = Paginator(users, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(request, "fss/users_admin.html", {
        "users": page_obj,
        "query": query,
    })


@csrf_exempt
def delete_user(request, user_id):
    if request.method == "POST":
        try:
            user = CustomUser.objects.get(id=user_id)
            user.delete()
            return JsonResponse({"success": True})
        except CustomUser.DoesNotExist:
            return JsonResponse({"success": False, "error": "User not found"})
    return JsonResponse({"success": False, "error": "Invalid request"})


def edit_user(request, user_id):
    user = get_object_or_404(CustomUser, pk=user_id)

    if request.method == 'POST':
        form = CustomUserEditForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Пользователь успешно обновлён.')
            return redirect('user_management')
    else:
        form = CustomUserEditForm(instance=user)

    return render(request, 'fss/edit_user.html', {'form': form, 'user': user})

@login_required
@require_POST
def mark_notifications_read(request):
    Notification.objects.filter(user=request.user, is_read=False).update(is_read=True)
    return JsonResponse({"success": True})

@require_POST
@login_required
def rate_suggestion(request):
    try:
        data = json.loads(request.body)
        suggestion_id = data.get('suggestion_id')
        rating = int(data.get('rating'))

        suggestion = Suggestion.objects.get(pk=suggestion_id)

        # Проверяем, голосовал ли уже пользователь
        existing_rating = SuggestionRating.objects.filter(user=request.user, suggestion=suggestion).first()
        if existing_rating:
            return JsonResponse({
                'success': False,
                'error': 'Вы уже голосовали за это предложение и не можете изменить оценку.'
            }, status=400)

        # Создаем новую оценку
        SuggestionRating.objects.create(
            user=request.user,
            suggestion=suggestion,
            rating=rating
        )

        # Считаем средний рейтинг и количество голосов
        agg = SuggestionRating.objects.filter(suggestion=suggestion).aggregate(
            avg=Avg('rating'),
            votes=Count('id')
        )
        avg_rating = agg['avg'] or 0.0
        votes = agg['votes'] or 0

        return JsonResponse({
            'success': True,
            'new_avg_rating': round(avg_rating, 2),
            'votes': votes,
        })

    except Suggestion.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Предложение не найдено.'
        }, status=404)

    except (ValueError, TypeError):
        return JsonResponse({
            'success': False,
            'error': 'Некорректные данные.'
        }, status=400)

    except Exception as e:
        return JsonResponse({'success': False, 'error': 'Произошла ошибка на сервере.'}, status=500)

def add_user(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            form.save()  # автоматически хеширует пароль и сохраняет пользователя
            messages.success(request, 'Пользователь успешно добавлен.')
            return redirect('user_management')  # убедись, что такой URL-нейм существует
    else:
        form = CustomUserCreationForm()

    return render(request, 'fss/add_user.html', {'form': form})